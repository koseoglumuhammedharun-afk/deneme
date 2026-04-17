# -*- coding: utf-8 -*-
"""
Rapor oluşturma (Excel ve JSON dışa aktarma)
"""

import json
import logging
import shutil
from datetime import datetime
from pathlib import Path
from typing import Dict, Optional

try:
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    pd = None

import config
from .utils import ensure_output_directory

logger = logging.getLogger(__name__)


class ReportGenerator:
    """Excel ve JSON raporları oluştur"""

    @staticmethod
    def create_excel_report(analysis_data: list, output_filename: Optional[str] = None) -> Optional[str]:
        """
        Analiz sonuçlarıyla Excel raporu oluştur

        Args:
            analysis_data: Analiz sonuç sözlükleri listesi
            output_filename: Özel çıktı dosya adı veya otomatik oluşturulsun

        Returns:
            Oluşturulan dosyanın yolu veya başarısız olursa None
        """
        try:
            if pd is None:
                logger.error("pandas yüklenmedi")
                return None

            ensure_output_directory()

            if isinstance(analysis_data, dict):
                analysis_data = [analysis_data]

            excel_data = []
            for data in analysis_data:
                row = {
                    "Dosya Adı": data.get("filename", "N/A"),
                    "Dosya Türü": data.get("file_type", "N/A"),
                    "Tespit Durumu": "Evet" if data.get("detected", False) else "Hayır",
                    "Silah Türü": data.get("weapon_type", data.get("class_name", "N/A")),
                    "Sınıf ID": data.get("class_id", "N/A"),
                    "Güven Puanı": round(data.get("confidence", 0), 4),
                    "Tespit Zamanı (DD:SS)": data.get("time_in_video", "N/A"),
                    "Çekim Tarihi": data.get("capture_date", "N/A"),
                    "Çekim Saati": data.get("capture_time", "N/A"),
                    "Analiz Tarihi": data.get("analysis_date", "N/A"),
                    "Analiz Saati": data.get("analysis_time", "N/A"),
                    "Tahmini Mesafe (m)": data.get("distance_m", "N/A"),
                    "GPS Enlemi": data.get("gps_latitude", "N/A"),
                    "GPS Boylamı": data.get("gps_longitude", "N/A"),
                }
                excel_data.append(row)

            df = pd.DataFrame(excel_data)

            if output_filename is None:
                timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                output_filename = f"analysis_report_{timestamp}.xlsx"

            output_path = config.OUTPUTS_DIR / output_filename

            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=config.EXCEL_SHEET_NAME, index=False)

                worksheet = writer.sheets[config.EXCEL_SHEET_NAME]

                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")

                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                widths = {
                    "A": 25,
                    "B": 14,
                    "C": 16,
                    "D": 18,
                    "E": 12,
                    "F": 14,
                    "G": 18,
                    "H": 15,
                    "I": 15,
                    "J": 15,
                    "K": 15,
                    "L": 20,
                    "M": 18,
                    "N": 18,
                }
                for col, width in widths.items():
                    worksheet.column_dimensions[col].width = width

                for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        cell.border = Border(
                            left=Side(style="thin"),
                            right=Side(style="thin"),
                            top=Side(style="thin"),
                            bottom=Side(style="thin"),
                        )

                for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1, min_col=3, max_col=3):
                    for cell in row:
                        if cell.value == "Evet":
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif cell.value == "Hayır":
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            logger.info(f"Excel raporu oluşturuldu: {output_path}")
            return str(output_path)

        except Exception as e:
            logger.error(f"Excel raporu oluşturma hatası: {e}")
            return None

    @staticmethod
    def create_json_report(analysis_data: Dict, output_filename: Optional[str] = None) -> Optional[str]:
        """
        Analiz sonuçlarıyla JSON raporu oluştur

        Args:
            analysis_data: Analiz sonuç sözlüğü
            output_filename: Özel çıktı dosya adı veya otomatik oluşturulsun

        Returns:
            Oluşturulan dosyanın yolu veya başarısız olursa None
        """
        try:
            ensure_output_directory()

            json_output = {
                "detected": analysis_data.get("detected", False),
                "confidence": round(analysis_data.get("confidence", 0), 4),
                "weapon_type": analysis_data.get("weapon_type", analysis_data.get("class_name", None)),
                "class_id": analysis_data.get("class_id", None),
                "class_name": analysis_data.get("class_name", None),
                "time_in_video": analysis_data.get("time_in_video", None),
                "capture_date": analysis_data.get("capture_date", None),
                "capture_time": analysis_data.get("capture_time", None),
                "analysis_datetime": analysis_data.get(
                    "analysis_datetime",
                    datetime.now().isoformat()
                ),
                "distance_m": analysis_data.get("distance_m", None),
                "gps": {
                    "lat": analysis_data.get("gps_latitude", None),
                    "lon": analysis_data.get("gps_longitude", None),
                },
                "crop_image_path": analysis_data.get("crop_image_path", None),
                "metadata": {
                    "filename": analysis_data.get("filename", None),
                    "file_type": analysis_data.get("file_type", None),
                    "capture_source": analysis_data.get("capture_source", None),
                },
            }

            if output_filename is None:
                timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                output_filename = f"analysis_report_{timestamp}.json"

            output_path = config.OUTPUTS_DIR / output_filename

            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(json_output, f, indent=config.JSON_INDENT, ensure_ascii=False)

            logger.info(f"JSON raporu oluşturuldu: {output_path}")
            return str(output_path)

        except Exception as e:
            logger.error(f"JSON raporu oluşturma hatası: {e}")
            return None

    @staticmethod
    def export_to_excel(analysis_data, file_path: str) -> Optional[str]:
        """
        GUI'nin çağırdığı dışa aktarma metodu.
        create_excel_report ile dosyayı üretir, gerekirse hedef yola kopyalar.
        """
        try:
            target_path = Path(file_path)
            result_path = ReportGenerator.create_excel_report(analysis_data, target_path.name)

            if result_path is None:
                return None

            result_path = Path(result_path)

            if result_path.resolve() != target_path.resolve():
                target_path.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(result_path, target_path)

            logger.info(f"Excel dışa aktarıldı: {target_path}")
            return str(target_path)

        except Exception as e:
            logger.error(f"Excel dışa aktarma hatası: {e}")
            return None

    @staticmethod
    def export_to_json(analysis_data, file_path: str) -> Optional[str]:
        """
        GUI'nin çağırdığı dışa aktarma metodu.
        create_json_report ile dosyayı üretir, gerekirse hedef yola kopyalar.
        """
        try:
            target_path = Path(file_path)
            result_path = ReportGenerator.create_json_report(analysis_data, target_path.name)

            if result_path is None:
                return None

            result_path = Path(result_path)

            if result_path.resolve() != target_path.resolve():
                target_path.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(result_path, target_path)

            logger.info(f"JSON dışa aktarıldı: {target_path}")
            return str(target_path)

        except Exception as e:
            logger.error(f"JSON dışa aktarma hatası: {e}")
            return None

    @staticmethod
    def save_crop_image(crop_image, filename: Optional[str] = None) -> Optional[str]:
        """
        Kırpılmış tespit görüntüsünü kaydet

        Args:
            crop_image: Görüntü dizisi (numpy veya PIL)
            filename: Özel dosya adı veya otomatik oluşturulsun

        Returns:
            Kaydedilen görüntünün yolu veya başarısız olursa None
        """
        try:
            import cv2
            from PIL import Image

            ensure_output_directory()

            if filename is None:
                timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                filename = f"detection_crop_{timestamp}.jpg"

            output_path = config.OUTPUTS_DIR / filename

            if isinstance(crop_image, Image.Image):
                crop_image.save(output_path)
            else:
                cv2.imwrite(str(output_path), crop_image)

            logger.info(f"Kırpma görüntüsü kaydedildi: {output_path}")
            return str(output_path)

        except Exception as e:
            logger.error(f"Kırpma görüntüsü kaydetme hatası: {e}")
            return None